#! /usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')

print(wb.sheetnames)
print(wb.active)
print(sheet.max_row)
print(sheet.max_column)

print(' --- start of printing contents --- ')

for rowOfCellObjects in sheet['A1':'C3']:
    print(rowOfCellObjects)
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print(' --- EOD OF ROW --- ')
